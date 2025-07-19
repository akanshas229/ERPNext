# Copyright (c) 2025, pinkcity and contributors
# For license information, please see license.txt

import frappe, openpyxl, json, io
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

def execute(filters=None):
	if not filters:
		filters = {}

	columns = get_columns()
	data = get_data(filters)

	updated_data = update_data(data)

	return columns, updated_data

def update_data(data) :
	for row in data :
		if float(row.get("amount", 0) or 0) >= 15000 :
			row['amount'] = 15000
		row['pf'] = round(float(float(row.get("amount", 0) or 0) * (12 / 100 ) ))
		row['epsepf'] = round(float(float(row.get("pf", 0) or 0) - float( float(row.get("amount", 0) or 0) * (8.33 / 100 ) or 0 ) ))
		row['eps'] = round(float(float(row.get("amount", 0) or 0) * (8.33 / 100 ) ))
		if int(row.get("eps_scheme_not_applicable", 0) or 0) == 1 :
			row['epsepf'] = row["pf"]
			row['eps'] = 0

	for row in data :
		if float(row.get('gross_pay', 0) or 0) :
			pass
		else :
			row.amount = 0
			row.pf = 0
			row.eps = 0
			row.epsepf = 0
	return data

def get_columns():
	return [
		{"fieldname":"uan_number", "fieldtype":"Int", "label":"UAN", "width":130},
		{"fieldname":"employee_name", "fieldtype":"Data", "label":"Name of Member", "width":150},
		{"fieldname":"amount", "fieldtype":"Currency", "label":"Employee PF Earnings", "width":190},
		{"fieldname":"pf", "fieldtype":"Currency", "label":"Contribution EPF", "width":160},
		{"fieldname":"epsepf", "fieldtype":"Currency", "label":"Employer EPF Difference", "width":220},
		{"fieldname":"eps", "fieldtype":"Currency", "label":"Contribution Pension 8.33%", "width":220},
	]

def get_data(filters):
	conditions = get_conditions(filters)

	# query = f"""	SELECT
	# 					tss.uan_number,
	# 					tss.employee_name,
	# 					round(tss.gross_pay) gross_pay,
	# 					IF(tsd.amount<15000, round(tsd.amount),  round(15000)  ) amount,
	# 					IF(tsd.amount<15000, round(((tsd.amount*12)/100)),  round(((15000*12)/100))  ) pf,
	# 					IF(tsd.amount<15000, round(((tsd.amount*8.33)/100)),  round(((15000*8.33)/100))  ) eps,
	# 					IF(tsd.amount<15000,
	# 					( round(((tsd.amount*12)/100)) - round(((tsd.amount*8.33)/100) )  ),
	# 					( round(((15000*12)/100)) - round(((15000*8.33)/100) )  )  )  epsepf,
	# 					tss.eps_scheme_not_applicable
	# 				FROM `tabSalary Slip` tss
	# 				LEFT JOIN `tabSalary Detail` tsd ON tss.name = tsd.parent AND tsd.abbr = 'B'
	# 				WHERE tss.eps_scheme_not_applicable = 0
	# 					  AND tss.uan_number != ''
	# 					  {conditions}
	# 			UNION 
	# 				SELECT
	# 					tss.uan_number,
	# 					tss.employee_name,
	# 					round(tss.gross_pay) gross_pay,
	# 					IF(tsd.amount<15000, round(tsd.amount),  round(15000)  ) amount,
	# 					IF(tsd.amount<15000, round(((tsd.amount*12)/100)),  round(((15000*12)/100))  ) pf,
	# 					0 eps,
	# 					IF(amount<15000, round(((amount*12)/100)),  round(((15000*12)/100))  ) epsepf,
	# 					tss.eps_scheme_not_applicable
	# 				FROM `tabSalary Slip` tss
	# 				LEFT JOIN `tabSalary Detail` tsd ON tss.name = tsd.parent AND tsd.abbr = 'B'
	# 				WHERE tss.eps_scheme_not_applicable = 1
	# 					  AND tss.uan_number != ''
	# 					  {conditions}
	# 			ORDER BY uan_number ASC
			# """

	query = f"""SELECT
					tss.uan_number,
					tss.employee_name,
					round(tss.gross_pay) gross_pay,
					tsd.amount,
					tss.eps_scheme_not_applicable
				FROM `tabSalary Slip` tss
				LEFT JOIN `tabSalary Detail` tsd ON tss.name = tsd.parent AND tsd.abbr = 'B'
				WHERE LENGTH(tss.uan_number) > 4 AND tss.docstatus <= 1
						{conditions}
				ORDER BY uan_number ASC
			"""
	return frappe.db.sql(query, as_dict=1)




def get_all_data(filters):
	conditions = get_conditions(filters)

	query = f"""SELECT
					tss.uan_number,
					tss.employee_name,
					round(tss.gross_pay) gross_pay,
					tss.uan_number
				FROM `tabSalary Slip` tss
				WHERE tss.docstatus <= 1
					{conditions}
				ORDER BY uan_number ASC
			"""
	return frappe.db.sql(query, as_dict=1)

def get_conditions(filters):
	conditions = ""
	
	if filters.get("company"):
		conditions += " AND tss.company = '" + filters.get("company") +"' " 

	if filters.get("month"):
		month = [
			"Jan",
			"Feb",
			"Mar",
			"Apr",
			"May",
			"Jun",
			"Jul",
			"Aug",
			"Sep",
			"Oct",
			"Nov",
			"Dec",
		].index(filters["month"]) + 1
		conditions += f" AND MONTH(tss.start_date) = {month} "
	
	if filters.get("year"):
		conditions += f" AND YEAR(tss.start_date) =  " + filters.get("year") +" " 
	
	if filters.get("employee"):
		conditions += f" AND tss.employee =  '" + filters.get("employee") +"' " 
	

	return conditions





@frappe.whitelist()
def get_excel_data(filters):
	filters = json.loads(filters)
	# report_data = json.loads(data)
	file_name = "Provident Fund Statement"
	if filters.get("company"):
		file_name  += " - "+ filters.get("company") 
	if filters.get("month"):
		file_name  += " - "+ filters.get("month") 
	if filters.get("year"):
		file_name  += " - "+ filters.get("year") 
	data = ""

	return {
        "content": data,
        "filename": file_name,  
        "extension": 'xlsx',
        "filters":filters
    }


@frappe.whitelist()
def download_file():    

	data = frappe._dict(frappe.local.form_dict)
	filters =  json.loads(data["filters"])

	# workbook = openpyxl.Workbook(write_only=True)
	workbook = openpyxl.Workbook()
	current_sheet = workbook.active
	border_style_thin = Side(border_style='thin')
	border_style_thick = Side(border_style='thick')
    # bold_font = Font(bold=True)

	current_sheet.row_dimensions[1].height = 45
	current_sheet.column_dimensions['A'].width = 15
	current_sheet.column_dimensions['B'].width = 18
	current_sheet.column_dimensions['C'].width = 30
	current_sheet.column_dimensions['D'].width = 20
	current_sheet.column_dimensions['E'].width = 20
	current_sheet.column_dimensions['F'].width = 20
	current_sheet.column_dimensions['G'].width = 20

	update_border_font_align(current_sheet, 'A1', "S.No. \n(1)", border_style_thick)
	update_border_font_align(current_sheet, 'B1', "UAN \n(2)", border_style_thick) 
	update_border_font_align(current_sheet, 'C1', "Name of Member \n(3)", border_style_thick)
	update_border_font_align(current_sheet, 'D1', "Employee PF Earnings \n(4)", border_style_thick)
	update_border_font_align(current_sheet, 'E1', "Contribution EPF \n(5)", border_style_thick)
	update_border_font_align(current_sheet, 'F1', "Employer EPF Difference \n(6)", border_style_thick)
	update_border_font_align(current_sheet, 'G1', "Contribution Pension 8.33% \n(7)", border_style_thick)


	total_amount = 0
	total_pf = 0
	total_eps = 0
	total_epsepf = 0

	account_01 = 0
	account_02 = 0
	account_10 = 0
	pension_wages = 0
	all_total = 0

	total_eps_scheme_not_applicable = 0

	total_employee = 0
	excluded_employee = 0
	excluded_employee_gross = 0

	temp_data = get_data(filters)
	query_data = update_data(temp_data)

	temp_data2 = get_all_data(filters)
	for row2 in temp_data2:
		total_employee += 1
		if row2.get("uan_number", "") :
			pass
		else:
			excluded_employee += 1
			excluded_employee_gross += float(row2.get("gross_pay", 0) or 0)


	row_no = 2
	for row in query_data :
		update_border_font_align(current_sheet, f"A{row_no}", row_no - 1, border_style_thin, bold='no' )
		update_border_font_align(current_sheet, f"B{row_no}", row.get("uan_number", ""), border_style_thin, bold='no' )
		update_border_font_align(current_sheet, f"C{row_no}", row.get("employee_name", ""), border_style_thin, bold='no' )
		update_border_font_align(current_sheet, f"D{row_no}", format_number(row.get("amount", 0)), border_style_thin,  bold='no', number_format="0" )
		update_border_font_align(current_sheet, f"E{row_no}", format_number(row.get("pf", 0)), border_style_thin,  bold='no',  number_format="0" )
		update_border_font_align(current_sheet, f"F{row_no}", format_number(row.get("epsepf", 0)), border_style_thin,  bold='no',  number_format="0" )
		update_border_font_align(current_sheet, f"G{row_no}", format_number(row.get("eps", 0)), border_style_thin,  bold='no',  number_format="0" )
		if row.get("eps_scheme_not_applicable", 0) == 1 :
			total_eps_scheme_not_applicable += float(row.get("amount", 0) or 0)
		total_amount += float(row.get("amount", 0) or 0)
		total_pf += float(row.get("pf", 0) or 0)
		total_epsepf += float(row.get("epsepf", 0) or 0)
		total_eps += float(row.get("eps", 0) or 0)
		row_no += 1

	account_01 = float(total_pf or 0) + float(total_eps or 0)
	account_02 = round( float(total_amount or 0) * 0.5000/100 )
	account_10 = total_epsepf
	pension_wages = float(total_amount or 0) - float(total_eps_scheme_not_applicable or 0)
	all_total = float(total_pf or 0) + float(total_eps or 0) + float(total_epsepf or 0) + float(account_02 or 0) + float(account_02 or 0)

	current_sheet.merge_cells(f"A{row_no}:C{row_no}")
	update_border_font_align(current_sheet, f"A{row_no}", "Total", border_style_thin, bold='no', text_align="right")
	update_border_font_align(current_sheet, f"D{row_no}", format_number(total_amount), border_style_thin, bold="no", )
	update_border_font_align(current_sheet, f"E{row_no}", format_number(total_pf), border_style_thin, bold="no", )
	update_border_font_align(current_sheet, f"F{row_no}", format_number(total_epsepf), border_style_thin, bold="no", )
	update_border_font_align(current_sheet, f"G{row_no}", format_number(total_eps), border_style_thin, bold="no", )
		
	row_no += 1
	current_sheet.merge_cells(f"A{row_no}:C{row_no}")
	update_border_font_align(current_sheet, f"A{row_no}", "Account No:01", border_style_thin, bold='no', text_align="right")
	current_sheet.merge_cells(f"D{row_no}:E{row_no}")
	update_border_font_align(current_sheet, f"D{row_no}", "(Column Nos.5+6)", border_style_thin, text_align="left")
	update_border_font_align(current_sheet, f"F{row_no}", "=", border_style_thin, )
	update_border_font_align(current_sheet, f"G{row_no}", format_number(account_01), border_style_thin, number_format="0" )

	row_no += 1
	current_sheet.merge_cells(f"A{row_no}:C{row_no}")
	update_border_font_align(current_sheet, f"A{row_no}", "Account No:02", border_style_thin, bold='no', text_align="right")
	current_sheet.merge_cells(f"D{row_no}:E{row_no}")
	update_border_font_align(current_sheet, f"D{row_no}", "(0.50000% of Column Nos.4)", border_style_thin, text_align="left")
	update_border_font_align(current_sheet, f"F{row_no}", "=", border_style_thin, )
	update_border_font_align(current_sheet, f"G{row_no}", format_number(account_02), border_style_thin, number_format="0" )

	row_no += 1
	current_sheet.merge_cells(f"A{row_no}:C{row_no}")
	update_border_font_align(current_sheet, f"A{row_no}", "Account No:10", border_style_thin, bold='no', text_align="right")
	current_sheet.merge_cells(f"D{row_no}:E{row_no}")
	update_border_font_align(current_sheet, f"D{row_no}", "(Column Nos.7)", border_style_thin, text_align="left")
	update_border_font_align(current_sheet, f"F{row_no}", "=", border_style_thin, )
	update_border_font_align(current_sheet, f"G{row_no}", format_number(account_10), border_style_thin, number_format="0" )

	row_no += 1
	update_border_font_align(current_sheet, f"A{row_no}", "EDLI Wages :", border_style_thin, bold='no', text_align="left")
	update_border_font_align(current_sheet, f"B{row_no}", format_number(total_amount), border_style_thin, )
	update_border_font_align(current_sheet, f"C{row_no}", "Account No:21", border_style_thin, bold='no', text_align="right")
	current_sheet.merge_cells(f"D{row_no}:E{row_no}")
	update_border_font_align(current_sheet, f"D{row_no}", "EDLI Wages * 0.50000%", border_style_thin, text_align="left")
	update_border_font_align(current_sheet, f"F{row_no}", "=", border_style_thin, )
	update_border_font_align(current_sheet, f"G{row_no}", format_number(account_02), border_style_thin, number_format="0" )

	row_no += 1
	update_border_font_align(current_sheet, f"A{row_no}", "Pension Wages :", border_style_thin, bold='no', text_align="left")
	update_border_font_align(current_sheet, f"B{row_no}", format_number(pension_wages), border_style_thin, )
	update_border_font_align(current_sheet, f"C{row_no}", "Account No:22", border_style_thin, bold='no', text_align="right")
	current_sheet.merge_cells(f"D{row_no}:E{row_no}")
	update_border_font_align(current_sheet, f"D{row_no}", "EDLI Wages * 0.00000%", border_style_thin, text_align="left")
	update_border_font_align(current_sheet, f"F{row_no}", "=", border_style_thin, )
	update_border_font_align(current_sheet, f"G{row_no}", format_number(0), border_style_thin, number_format="0" )

	row_no += 1
	current_sheet.merge_cells(f"A{row_no}:C{row_no}")
	update_border_font_align(current_sheet, f"A{row_no}", "Total", border_style_thin, bold='no', text_align="right")
	current_sheet.merge_cells(f"D{row_no}:F{row_no}")
	update_border_font_align(current_sheet, f"D{row_no}", "", border_style_thin, )
	update_border_font_align(current_sheet, f"G{row_no}", format_number(all_total), border_style_thin, number_format="0" )

	row_no += 1
	current_sheet.merge_cells(f"A{row_no}:C{row_no}")
	update_border_font_align(current_sheet, f"A{row_no}", "Total No. of Employees in the Month :", border_style_thin, bold='no', text_align="left")
	update_border_font_align(current_sheet, f"D{row_no}", format_number(total_employee), border_style_thin, number_format="0" )

	row_no += 1
	current_sheet.merge_cells(f"A{row_no}:C{row_no}")
	update_border_font_align(current_sheet, f"A{row_no}", "No. of Excluded Employee :", border_style_thin, bold='no', text_align="left")
	update_border_font_align(current_sheet, f"D{row_no}", format_number(excluded_employee), border_style_thin, number_format="0" )

	row_no += 1
	current_sheet.merge_cells(f"A{row_no}:C{row_no}")
	update_border_font_align(current_sheet, f"A{row_no}", "Gross Wages of Excluded Employee :", border_style_thin, bold='no', text_align="left")
	update_border_font_align(current_sheet, f"D{row_no}", format_number(excluded_employee_gross), border_style_thin, number_format="0" )

						  

	byte_file = io.BytesIO()
	workbook.save(byte_file)

	frappe.response["type"] = "binary"
	frappe.response["filecontent"] = byte_file.getvalue()
	frappe.response["filename"] = f"{data['filename']}.{data['extension']}"


def set_border(border_type):
	return Border(top=border_type, left=border_type, right=border_type, bottom=border_type)

def update_border_font_align(sheet, index, name, border_type, bold='yes', number_format="General", text_align = "center") :
    sheet[index] = name
    if bold == 'yes':
        sheet[index].font = Font(bold=True)
    sheet[index].border = set_border(border_type)
    sheet[index].alignment = Alignment(horizontal=text_align, vertical='center', wrap_text=True)
    sheet[index].number_format = number_format
    # sheet[index].width = 200
    return sheet




def format_number(value) :
	# return value
    if float(value or 0):
        #   return "{:,.0f}0".format(value)
        #   return frappe.utils.fmt_money(value, precision=0, currency='INR')
          return frappe.utils.fmt_money(value, precision=0, )
    else :
          return 0
      

