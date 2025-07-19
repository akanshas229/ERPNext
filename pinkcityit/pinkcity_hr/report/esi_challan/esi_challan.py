# Copyright (c) 2025, pinkcity and contributors
# For license information, please see license.txt

import frappe, openpyxl, json, io
# from openpyxl.cell.rich_text import CellRichText
# from openpyxl.cell.text import RichText
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill


def execute(filters=None):
	columns = get_columns()
	data = get_data(filters)
	return columns, data




def get_data(filters):

	conditions = ""
	emp_conditions = ""
	month = 1
	year = 2025
	previous_month = 1
	previous_year = 2024
	
	if filters.get("company"):
		conditions += " AND tss.company = '" + filters.get("company") +"' " 
		emp_conditions += " AND te.company = '" + filters.get("company") +"' " 


	if filters.get("year"):
		year = int(filters.get("year") or 2025)
		previous_year = year
		conditions += f" AND YEAR(tss.start_date) =  " + filters.get("year") +" " 

	if filters.get("month"):
		month = [
			"Jan",
			"Feb",
			"Mar",
			"Apr",
			"May",
			"Jun",
			"Jul",
			"Aug",
			"Sep",
			"Oct",
			"Nov",
			"Dec",
		].index(filters["month"]) + 1
		conditions += f" AND MONTH(tss.start_date) = {month} "
		previous_month = int(month or 1) - 1
		if previous_month == 0:
			previous_month = 12
			previous_year = int(year or 2025) - 1
		emp_conditions += f" AND ( ( MONTH(esic_exit_date) = {previous_month} AND YEAR(esic_exit_date) = {previous_year} ) )" 
		
	
	
	if filters.get("employee"):
		conditions += f" AND tss.employee =  '" + filters.get("employee") +"' " 
		emp_conditions += f" AND te.employee =  '" + filters.get("employee") +"' " 


	query = f"""
				SELECT 
					tss.esic_no,
					tss.employee_name,
					ROUND(tss.payment_days) AS payment_days,
					ROUND(
						tss.gross_pay - IFNULL(
												(SELECT td.amount 
												 FROM `tabSalary Detail` td 
											     WHERE td.parent = `tabSalary Detail`.parent AND td.abbr = 'WA'), 
										0)
					) AS esi_earnings,
					(IF(payment_days > 0, 0, 1)) AS workings_day1,
					tss.esic_exit_date AS last_working_day
				FROM `tabSalary Slip` tss
				LEFT JOIN `tabSalary Detail` ON tss.name = `tabSalary Detail`.parent AND `tabSalary Detail`.abbr = 'B'
				WHERE tss.esic_exit_date IS NULL
					  AND tss.esic_no IS NOT NULL
					  AND (SELECT tsd.amount  FROM `tabSalary Detail` tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'ESI') > 0
					  AND payment_days IS NOT NULL
					  {conditions}
			UNION 
				SELECT
					te.esic_no,
                    te.employee_name,
                    0 payment_days,
                    0 esi_earnings,
                    2 workings_day1,
                    DATE_FORMAT(te.esic_exit_date, '%d-%m-%Y') as last_working_day
				FROM tabEmployee te
				WHERE te.esic_no IS NOT NULL
					  {emp_conditions}
	
				"""


	return frappe.db.sql(query, as_dict=1,)



def get_columns():
	return [
		{"fieldname":"esic_no", "fieldtype":"Data", "label":"IP Number (10 Digits)", "width":250},
		{"fieldname":"employee_name", "fieldtype":"Data", "label":"IP Name (Only alphabets and space)", "width":250},
		{"fieldname":"payment_days", "fieldtype":"Data", "label":"No of Days for which wages paid payable during the month", "width":250},
		{"fieldname":"esi_earnings", "fieldtype":"Data", "label":"Total Monthly Wages", "width":250},
		{"fieldname":"workings_day1", "fieldtype":"Data", "label":"Reason Code for Zero workings days (numeric only; provide 0 for all other reasons- Click on the link for referencel)", "width":250},
		{"fieldname":"last_working_day", "fieldtype":"Date", "label":"Last Working Day (Format DD/MM/YYYY or DD-MM-YYYY)", "width":250},

	]




@frappe.whitelist()
def get_excel_data(filters):
	filters = json.loads(filters)
	# report_data = json.loads(data)
	file_name = "ESI Statement"
	if filters.get("company"):
		file_name  += " - "+ filters.get("company") 
	if filters.get("month"):
		file_name  += " - "+ filters.get("month") 
	if filters.get("year"):
		file_name  += " - "+ filters.get("year") 
	data = ""

	return {
        "content": data,
        "filename": file_name,  
        "extension": 'xls',
        "filters":filters
    }


@frappe.whitelist()
def download_file():    

	data = frappe._dict(frappe.local.form_dict)
	filters =  json.loads(data["filters"])

	# workbook = openpyxl.Workbook(write_only=True)
	workbook = openpyxl.Workbook()
	current_sheet = workbook.active
	border_style_thin = Side(border_style='thin')
	border_style_thick = Side(border_style='thick')
	red_color = Font(color="00FF0000")


	current_sheet.row_dimensions[1].height = 70
	current_sheet.column_dimensions['A'].width = 17
	current_sheet.column_dimensions['B'].width = 25
	current_sheet.column_dimensions['C'].width = 30
	current_sheet.column_dimensions['D'].width = 15
	current_sheet.column_dimensions['E'].width = 30
	current_sheet.column_dimensions['F'].width = 20

	# first_string  = CellRichText("IP Number " + InlineFont(red_color, "(10 Digits)") )
	# first_string  = "IP Number " + InlineFont(red_color, "(10 Digits)") 
	# first_string  =  InlineFont(red_color, "(10 Digits)") 
	second_string  = ""
	third_string  = ""
	forth_string  = ""

	update_border_font_align(current_sheet, 'A1', "IP Number (10 Digits)", border_style_thick)
	update_border_font_align(current_sheet, 'B1', "IP Name (Only alphabets and space)", border_style_thick) 
	update_border_font_align(current_sheet, 'C1', "No of Days for which wages paid payable during the month", border_style_thick)
	update_border_font_align(current_sheet, 'D1', "Total Monthly Wages", border_style_thick)
	update_border_font_align(current_sheet, 'E1', "Reason Code for Zero workings days (numeric only; provide 0 for all other reasons- Click on the link for referencel)", border_style_thick)
	update_border_font_align(current_sheet, 'F1', "Last Working Day (Format DD/MM/YYYY or DD-MM-YYYY)", border_style_thick)

	query_data = get_data(filters)
	row_no = 2
	for row in query_data :
		update_border_font_align(current_sheet, f"A{row_no}", row.get("esic_no", ""), border_style_thin, 'no' )
		update_border_font_align(current_sheet, f"B{row_no}", row.get("employee_name", ""), border_style_thin, 'no' )
		update_border_font_align(current_sheet, f"C{row_no}", row.get("payment_days", ""), border_style_thin, 'no',  number_format="0" )
		update_border_font_align(current_sheet, f"D{row_no}", format_number(row.get("esi_earnings", "")), border_style_thin,  'no', number_format="0" )
		update_border_font_align(current_sheet, f"E{row_no}", format_number(row.get("workings_day1", "")), border_style_thin,  'no',  number_format="0" )
		update_border_font_align(current_sheet, f"F{row_no}", row.get("last_working_day", ""), border_style_thin,  'no',  number_format="dd-mm-yy")
		row_no += 1 
	

	sheet_name = "Instructions & Reason Codes"
	second_sheet = workbook.create_sheet(sheet_name)
	# second_sheet = workbook.active

	# second_sheet.row_dimensions[1].height = 70
	second_sheet.column_dimensions['A'].width = 35
	second_sheet.column_dimensions['B'].width = 25
	second_sheet.column_dimensions['C'].width = 90

	update_border_font_align(second_sheet, 'A1', "Reason", border_style_thick)
	update_border_font_align(second_sheet, 'B1', "Code", border_style_thick)
	update_border_font_align(second_sheet, 'C1', "Code", border_style_thick)

	update_border_font_align(second_sheet, 'A2', "Without Reason", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'B2', "0", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'C2', "Leave last working day as blank", border_style_thin, 'no')

	update_border_font_align(second_sheet, 'A3', "On Leave", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'B3', "1", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'C3', "Leave last working day as blank", border_style_thin, 'no')

	update_border_font_align(second_sheet, 'A4', "Left Service", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'B4', "2", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'C4', "Please provide last working day (dd/mm/yyyy). IP will not appear from next wage period", border_style_thin, 'no')

	update_border_font_align(second_sheet, 'A5', "Retired", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'B5', "3", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'C5', "Please provide last working day (dd/mm/yyyy). IP will not appear from next wage period", border_style_thin, 'no')

	second_sheet.row_dimensions[6].height = 50
	update_border_font_align(second_sheet, 'A6', "Out of Coverage", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'B6', "4", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'C6', "Please provide last working day (dd/mm/yyyy). IP will not appear from next contribution period. This option is valid only if Wage Period is April/October. In case any other month then IP will continue to appear in the list", border_style_thin, 'no')

	update_border_font_align(second_sheet, 'A7', "Expired", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'B7', "5", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'C7', "Please provide last working day (dd/mm/yyyy). IP will not appear from next wage period", border_style_thin, 'no')

	update_border_font_align(second_sheet, 'A8', "Non Implemented area", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'B8', "6", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'C8', "Please provide last working day (dd/mm/yyyy).", border_style_thin, 'no')

	update_border_font_align(second_sheet, 'A9', "Compliance by Immediate Employer", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'B9', "7", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'C9', "Leave last working day as blank", border_style_thin, 'no')

	update_border_font_align(second_sheet, 'A10', "Suspension of work", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'B10', "8", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'C10', "Leave last working day as blank", border_style_thin, 'no')

	update_border_font_align(second_sheet, 'A11', "Strike/Lockout", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'B11', "9", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'C11', "Leave last working day as blank", border_style_thin, 'no')

	update_border_font_align(second_sheet, 'A12', "Retrenchment", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'B12', "10", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'C12', "Please provide last working day (dd/mm/yyyy). IP will not appear from next wage period", border_style_thin, 'no')

	update_border_font_align(second_sheet, 'A13', "No Work", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'B13', "11", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'C13', "Leave last working day as blank", border_style_thin, 'no')

	update_border_font_align(second_sheet, 'A14', "Doesn't Belong To This Employer", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'B14', "12", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'C14', "Leave last working day as blank", border_style_thin, 'no')

	update_border_font_align(second_sheet, 'A15', "Duplicate IP", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'B15', "13", border_style_thin, 'no')
	update_border_font_align(second_sheet, 'C15', "Leave last working day as blank", border_style_thin, 'no')
	
	# second_sheet.merge_cells(f"A18:C18")
	for x in range(18, 37) :
		second_sheet.merge_cells(f"A{x}:C{x}")

	second_sheet['A18'] = "Click Here to Go back to Data Entry Page"
	second_sheet.row_dimensions[18].height = 25
	second_sheet['A18'].font = Font(size=20, underline="single", color="00FF0000", bold=True)

	second_sheet.row_dimensions[19].height = 22
	second_sheet['A19'] = "Instructions to fill in the excel file:"
	second_sheet['A19'].font = Font(size=18, bold=True)

	second_sheet.row_dimensions[20].height = 30
	second_sheet['A20'] = "1. Enter the IP number, IP name, No. of Days, Total Monthly Wages, Reason for 0 wages, (If Wages '0') and Last Working Day( only if employee has left service, Retired, Out of coverage, Expired, Non-Implemented area or Retrenchment. For other reasons, last working day must be left BLANK)."
	second_sheet['A20'].alignment = Alignment(vertical='center', wrap_text=True)

	second_sheet['A21'] = "2. Number of days must me a whole number. Fractions should be rounded up to next higher whole number/integer."

	second_sheet['A22'] = "3. Excel sheet upload will lead to successful transaction only when all the Employees (who are currently mapped in the system) details are entered perfectly in the excel sheet."

	second_sheet['A23'] = "4. Reasons are to be assigned numeric code and date has to be provided as mentioned in the table above."

	second_sheet.row_dimensions[24].height = 30
	second_sheet['A24'] = "5. Once 0 wages given and last working day is mentioned as in reason codes (2,3,4,5,10) IP will be removed from the employer's record. Subsequent months will not have this IP listed under the employer. Last working day should be mentioned only if 'Number of days wages paid/payable' is '0'."
	second_sheet['A24'].alignment = Alignment(vertical='center', wrap_text=True)

	second_sheet['A25'] = "6. In case IP has worked for part of the month(i.e. atleast 1 day wage is paid/payable) and left in between of the month, then last working day shouldn't be mentioned."

	second_sheet['A26'] = "7. Calculations – IP Contribution and Employer contribution calculation will be automatically done by the system."

	second_sheet.row_dimensions[27].height = 30
	second_sheet['A27'] = "8. Date column format is dd/mm/yyyy or dd-mm-yyyy. Pad single digit dates with 0. Eg:- 2/5/2010 or 2-May-2010 is NOT acceptable. Correct format is 02/05/2010 or 02-05-2010."
	second_sheet['A27'].alignment = Alignment(vertical='center', wrap_text=True)

	second_sheet['A28'] = "9. Excel file should be saved in .xls format (Excel 97-2003)."

	second_sheet['A29'] = "10. Note that all the column including date column should be in 'Text' format."

	second_sheet['A30'] = "10a. To convert all columns to text,"

	second_sheet.row_dimensions[31].height = 30
	second_sheet['A31'] = "a. Select column A; Click Data in Menu Bar on top; Select Text to Columns ; Click Next (keep default selection of Delimited); Click Next (keep default selection of Tab); Select TEXT; Click FINISH. Excel 97 – 2003 as well have TEXT to COLUMN conversion facility"
	second_sheet['A31'].alignment = Alignment(vertical='center', wrap_text=True)

	second_sheet['A32'] = "b. Repeat the above step for each of the 6 columns. (Columns A – F )"

	second_sheet.row_dimensions[33].height = 30
	second_sheet['A33'] = "10b. Another method that can be used to text conversion is – copy the column with data and paste it in NOTEPAD. Select the column (in excel) and convert to text. Copy the data back from notepad to excel"
	second_sheet['A33'].alignment = Alignment(vertical='center', wrap_text=True)

	second_sheet.row_dimensions[34].height = 30
	second_sheet['A34'] = "11. If problem continues while upload, download a fresh template by clicking 'Sample MC Excel Template'. Then copy the data area from Step 8a.a – eg: copy Cell A2 to F8 (if there is data in 8 rows); Paste it in cell A2 in the fresh template. Upload it"
	second_sheet['A34'].alignment = Alignment(vertical='center', wrap_text=True)

	second_sheet.row_dimensions[36].height = 30
	second_sheet['A36'] = "Note : If problem continues while upload, download a fresh template by clicking 'Sample MC Excel Template'. Then copy the data area from Step 8a.a – eg: copy Cell A2 to F8 (if there is data in 8 rows); Paste it in cell A2 in the fresh template. Upload it"
	second_sheet['A36'].alignment = Alignment(vertical='center', wrap_text=True)

	second_sheet['A37'] = "1. Mozilla Firefox 3.5.11 : From Menu Bar, select Tools -> Options -> Content -> Uncheck (remove tick mark) 'Block Popup Windows'. Click OK"

	second_sheet['A38'] = "2. IE 7.0 : From Menu Bar, select Tools -> Pop up Blocker -> Turn Off Pop up Blocker"

	byte_file = io.BytesIO()
	workbook.save(byte_file)

	frappe.response["type"] = "binary"
	frappe.response["filecontent"] = byte_file.getvalue()
	frappe.response["filename"] = f"{data['filename']}.{data['extension']}"


def set_border(border_type):
	return Border(top=border_type, left=border_type, right=border_type, bottom=border_type)

def update_border_font_align(sheet, index, name, border_type, bold='yes', number_format="General") :
    sheet[index] = name
    if bold == 'yes':
        sheet[index].font = Font(bold=True)
    sheet[index].border = set_border(border_type)
    sheet[index].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet[index].number_format = number_format
    # sheet[index].width = 200
    return sheet




def format_number(value) :
     if float(value or 0):
        #   return "{:,.0f}0".format(value)
        #   return frappe.utils.fmt_money(value, precision=0, currency='INR')
          return frappe.utils.fmt_money(value, precision=0, )
     else :
          return 0
      
