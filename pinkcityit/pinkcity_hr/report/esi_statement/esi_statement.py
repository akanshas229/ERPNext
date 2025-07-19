# Copyright (c) 2025, Frappe Technologies Pvt. Ltd. and contributors
# For license information, please see license.txt

import frappe, openpyxl, json, io
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

def execute(filters=None):
	columns = get_columns()
	data = get_data(filters)


	report_summary = get_report_summary(data)

	return columns, data, None, None, report_summary


def get_columns():
	return [
		{"fieldname":"esic_no", "fieldtype":"Data", "label":"ESI NO.", "width":150},
		{"fieldname":"employee_name", "fieldtype":"Data", "label":"Name of Member", "width":250},
		{"fieldname":"payment_days", "fieldtype":"Data", "label":"Days Worked", "width":150},
		{"fieldname":"esi_earnings", "fieldtype":"Currency", "label":"ESI Earnings", "width":180},
		{"fieldname":"esi_contribution", "fieldtype":"Currency", "label":"ESI Contribution", "width":220},
		
	]

def get_data(filters):
	conditions = get_conditions(filters)

	query = f"""SELECT 
					tss.esic_no,
					tss.employee_name,
					ROUND(tss.payment_days) AS payment_days,
					ROUND(tss.gross_pay - IFNULL((
												SELECT tsd3.amount 
												FROM `tabSalary Detail` tsd3 
												WHERE tsd3.parent = tsd.parent AND tsd3.abbr = 'WA'
										), 0)) AS esi_earnings,
					ROUND(tsd.amount) AS esi_contribution
				FROM `tabSalary Slip` tss
				LEFT JOIN `tabSalary Detail` tsd ON tss.name = tsd.parent AND tsd.abbr = 'ESI'
				WHERE (
						SELECT amount 
						FROM `tabSalary Detail` tsd2 
						WHERE tsd2.parent = tss.name AND tsd2.abbr = 'ESI'
					) > 0
					AND tss.esic_exit_date IS NULL
					AND tss.esic_no IS NOT NULL
					{conditions}
				"""


	return frappe.db.sql(query, as_dict=1,)

def get_conditions(filters):
	conditions = ""
	
	if filters.get("company"):
		conditions += " AND tss.company = '" + filters.get("company") +"' " 

	if filters.get("month"):
		month = [
			"Jan",
			"Feb",
			"Mar",
			"Apr",
			"May",
			"Jun",
			"Jul",
			"Aug",
			"Sep",
			"Oct",
			"Nov",
			"Dec",
		].index(filters["month"]) + 1
		conditions += f" AND MONTH(ss.start_date) = {month} "
	
	if filters.get("year"):
		conditions += f" AND YEAR(ss.start_date) =  " + filters.get("year") +" " 
	
	if filters.get("employee"):
		conditions += f" AND ss.employee =  '" + filters.get("employee") +"' " 
	

	return conditions


def get_report_summary_data(data):
	if not data:
		return None
	
	employee_contribution = 0
	employer_contribution = 0
	total_esi_earnings = 0
	total_esi_contribution = 0
	total = 0

	for row in data :
		employee_contribution += float(row.get("esi_contribution", 0) or 0)
		total_esi_contribution += float(row.get("esi_contribution", 0) or 0)
		total_esi_earnings += float(row.get("esi_earnings", 0) or 0)
	employer_contribution = round(total_esi_earnings * 3.25 / 100 )
	total = employee_contribution + employer_contribution

	return {"employee_contribution": employee_contribution, "employer_contribution": employer_contribution, "total": total, "total_esi_contribution": total_esi_contribution, "total_esi_earnings": total_esi_earnings}


def get_report_summary(data):
	temp_data = get_report_summary_data(data)
	employee_contribution = temp_data['employee_contribution']
	employer_contribution = temp_data['employer_contribution']
	total = temp_data['total']

	return [
		{
			"value": employee_contribution,
			"indicator": "Blue" if employee_contribution > 50 else "Red",
			"label": "Employee Contribution",
			"datatype": "Currency",
		},
		{
			"value": employer_contribution,
			"indicator": "Blue",
			"label": "Employer Contribution",
			"datatype": "Currency",
		},
		{
			"value": total,
			"indicator": "Green",
			"label": "Total",
			"datatype": "Currency",
		},
	]



@frappe.whitelist()
def get_excel_data(filters):
	filters = json.loads(filters)
	# report_data = json.loads(data)
	file_name = "ESI Statement"
	if filters.get("company"):
		file_name  += " - "+ filters.get("company") 
	if filters.get("month"):
		file_name  += " - "+ filters.get("month") 
	if filters.get("year"):
		file_name  += " - "+ filters.get("year") 
	data = ""

	return {
        "content": data,
        "filename": file_name,  
        "extension": 'xlsx',
        "filters":filters
    }


@frappe.whitelist()
def download_file():    

	data = frappe._dict(frappe.local.form_dict)
	filters =  json.loads(data["filters"])

	# workbook = openpyxl.Workbook(write_only=True)
	workbook = openpyxl.Workbook()
	current_sheet = workbook.active
	border_style_thin = Side(border_style='thin')
	border_style_thick = Side(border_style='thick')
    # bold_font = Font(bold=True)

	current_sheet.column_dimensions['A'].width = 10
	current_sheet.column_dimensions['B'].width = 15
	current_sheet.column_dimensions['C'].width = 30
	current_sheet.column_dimensions['D'].width = 15
	current_sheet.column_dimensions['E'].width = 15
	current_sheet.column_dimensions['F'].width = 20

	update_border_font_align(current_sheet, 'A1', "S.No.", border_style_thick)
	update_border_font_align(current_sheet, 'B1', "ESI NO.", border_style_thick) 
	update_border_font_align(current_sheet, 'C1', "Name of Member", border_style_thick)
	update_border_font_align(current_sheet, 'D1', "Days Worked", border_style_thick)
	update_border_font_align(current_sheet, 'E1', "ESI Earnings", border_style_thick)
	update_border_font_align(current_sheet, 'F1', "ESI Contribution", border_style_thick)

	query_data = get_data(filters)
	temp_data = get_report_summary_data(query_data)
	row_no = 2
	for row in query_data :
		update_border_font_align(current_sheet, f"A{row_no}", row_no - 1, border_style_thin, 'no' )
		update_border_font_align(current_sheet, f"B{row_no}", row.get("esic_no", ""), border_style_thin, 'no' )
		update_border_font_align(current_sheet, f"C{row_no}", row.get("employee_name", ""), border_style_thin, 'no' )
		update_border_font_align(current_sheet, f"D{row_no}", row.get("payment_days", ""), border_style_thin, 'no',  number_format="0" )
		update_border_font_align(current_sheet, f"E{row_no}", format_number(row.get("esi_earnings", "")), border_style_thin,  'no', number_format="0" )
		update_border_font_align(current_sheet, f"F{row_no}", format_number(row.get("esi_contribution", "")), border_style_thin,  'no',  number_format="0" )
		row_no += 1
	current_sheet.merge_cells(f"A{row_no}:C{row_no}")
	update_border_font_align(current_sheet, f"A{row_no}", "Total", border_style_thin)
	update_border_font_align(current_sheet, f"D{row_no}", "", border_style_thin)
	update_border_font_align(current_sheet, f"E{row_no}", format_number(temp_data['total_esi_earnings']), border_style_thin)
	update_border_font_align(current_sheet, f"F{row_no}", format_number(temp_data['total_esi_contribution']), border_style_thin)
		
	row_no += 3
	current_sheet.merge_cells(f"A{row_no}:C{row_no}")
	update_border_font_align(current_sheet, f"A{row_no}", "Employee Contribution", border_style_thin)
	update_border_font_align(current_sheet, f"D{row_no}", format_number(temp_data['employee_contribution']), border_style_thin)
	
	row_no += 1
	current_sheet.merge_cells(f"A{row_no}:C{row_no}")
	update_border_font_align(current_sheet, f"A{row_no}", "Employer Contribution", border_style_thin)
	update_border_font_align(current_sheet, f"D{row_no}", format_number(temp_data['employer_contribution']), border_style_thin)
	
	row_no += 1
	current_sheet.merge_cells(f"A{row_no}:C{row_no}")
	update_border_font_align(current_sheet, f"A{row_no}", "Total", border_style_thin)
	update_border_font_align(current_sheet, f"D{row_no}", format_number(temp_data['total']), border_style_thin)
						  

	byte_file = io.BytesIO()
	workbook.save(byte_file)

	frappe.response["type"] = "binary"
	frappe.response["filecontent"] = byte_file.getvalue()
	frappe.response["filename"] = f"{data['filename']}.{data['extension']}"


def set_border(border_type):
	return Border(top=border_type, left=border_type, right=border_type, bottom=border_type)

def update_border_font_align(sheet, index, name, border_type, bold='yes', number_format="General") :
    sheet[index] = name
    if bold == 'yes':
        sheet[index].font = Font(bold=True)
    sheet[index].border = set_border(border_type)
    sheet[index].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet[index].number_format = number_format
    # sheet[index].width = 200
    return sheet




def format_number(value) :
     if float(value or 0):
        #   return "{:,.0f}0".format(value)
        #   return frappe.utils.fmt_money(value, precision=0, currency='INR')
          return frappe.utils.fmt_money(value, precision=0, )
     else :
          return 0
      
