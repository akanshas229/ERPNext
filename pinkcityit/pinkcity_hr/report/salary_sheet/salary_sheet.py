# Copyright (c) 2025, pinkcity and contributors
# For license information, please see license.txt

import frappe, openpyxl, json, io
# import openpyxl.styles
# import openpyxl.workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
# from frappe.utils import flt


def execute(filters):
    # headings = get_heading()
    columns = get_columns(filters)
    data = get_data(filters)

    temp_data = update_total_data(data, filters)
    report_data = temp_data['report_data']
    hdfc_total = temp_data['hdfc_total']
    other_total = temp_data['other_total']
    grand_total = temp_data['grand_total']

    # temp_data_1 = {}
    # temp_data_2 = {}

    # hdfc_total = 0
    # other_total = 0
    
    # for row in data :
    #         if row.get("occupation", False) :
    #             if row.get("occupation") == "Worker" :
    #                 update_data(row, 2, temp_data_2, "Worker")
    #             elif row.get("occupation") == "Staff"  :
    #                 update_data(row, 1, temp_data_1, "Staff")
    #             else :
    #                 update_data(row, 2, temp_data_2, "Worker")
    #         else :
    #             update_data(row, 2, temp_data_2, "Worker")


    #         if row.get("bank_name_new", False) :
    #             if row.get("bank_name_new") == "HDFC Bank" :
    #                 hdfc_total += float(row.get("net_pay", 0) or 0)
    #             else :
    #                 other_total += float(row.get("net_pay", 0) or 0)
    #         else :
    #             other_total += float(row.get("net_pay", 0) or 0)


    # report_data.append(temp_data_1)
    # report_data.append(temp_data_2)
    # total_data = {}


    # for key in temp_data_1:
    #     if key == "sr_no" :
    #          continue
        
    #     total_data[key] = temp_data_1.get(key, 0) + temp_data_2.get(key, 0)

    #     if key == "occupation" : 
    #         total_data[key] = "<b>Total</b>"
        


    # report_data.append(total_data)

    # grand_total = hdfc_total + other_total

    report_summary = get_report_summary(hdfc_total, other_total, grand_total)

    chart = get_chart_data(hdfc_total, other_total)


    return columns, report_data, None, chart, report_summary

def update_data(data_value, sr_no, temp_data, occupation) :
      temp_data['sr_no'] = sr_no
      temp_data['occupation'] = occupation
      temp_data['gross'] = float(temp_data.get("gross", 0) or 0) + float(data_value.get("gross_monthly_salary", 0) or 0)
      temp_data['total_employee'] = float(temp_data.get("total_employee", 0) or 0) + 1

      temp_data['basic'] = float(temp_data.get("basic", 0) or 0) + float(data_value.get("basic", 0) or 0)
      temp_data['city_compensatory_allowance'] = float(temp_data.get("city_compensatory_allowance", 0) or 0) + float(data_value.get("city_compensatory_allowance", 0) or 0)
      temp_data['washing_allowance'] = float(temp_data.get("washing_allowance", 0) or 0) + float(data_value.get("washing_allowance", 0) or 0)
      temp_data['house_rent_allowance'] = float(temp_data.get("house_rent_allowance", 0) or 0) + float(data_value.get("house_rent_allowance", 0) or 0)
      temp_data['overtime'] = float(temp_data.get("overtime", 0) or 0) + float(data_value.get("overtime", 0) or 0)
      temp_data['arrear'] = float(temp_data.get("arrear", 0) or 0) + float(data_value.get("arrear", 0) or 0)
      temp_data['incentive_pay'] = float(temp_data.get("incentive_pay", 0) or 0) + float(data_value.get("incentive_pay", 0) or 0)
      temp_data['leave_encashment'] = float(temp_data.get("leave_encashment", 0) or 0) + float(data_value.get("leave_encashment", 0) or 0)
      temp_data['travelling_allowance'] = float(temp_data.get("travelling_allowance", 0) or 0) + float(data_value.get("travelling_allowance", 0) or 0)
      temp_data['abry_scheme'] = float(temp_data.get("abry_scheme", 0) or 0) + float(data_value.get("abry_scheme", 0) or 0)

      temp_data['employee_contribution_pf'] = float(temp_data.get("employee_contribution_pf", 0) or 0) + float(data_value.get("employee_contribution_pf", 0) or 0)
      temp_data['employee_contribution_esi'] = float(temp_data.get("employee_contribution_esi", 0) or 0) + float(data_value.get("employee_contribution_esi", 0) or 0)
      temp_data['late_hours_deduction'] = float(temp_data.get("late_hours_deduction", 0) or 0) + float(data_value.get("late_hours_deduction", 0) or 0)
      temp_data['advance_deduction'] = float(temp_data.get("advance_deduction", 0) or 0) + float(data_value.get("advance_deduction", 0) or 0)
      temp_data['tax_deducted_at_source'] = float(temp_data.get("tax_deducted_at_source", 0) or 0) + float(data_value.get("tax_deducted_at_source", 0) or 0)
      temp_data['professional_tax'] = float(temp_data.get("professional_tax", 0) or 0) + float(data_value.get("professional_tax", 0) or 0)
      temp_data['income_tax'] = float(temp_data.get("income_tax", 0) or 0) + float(data_value.get("income_tax", 0) or 0)
      temp_data['gmi'] = float(temp_data.get("gmi", 0) or 0) + float(data_value.get("gmi", 0) or 0)


      temp_data['earned_gross'] = ( float(temp_data.get("basic", 0) or 0) + 
                                    float(temp_data.get("city_compensatory_allowance", 0) or 0) + 
                                    float(temp_data.get("washing_allowance", 0) or 0) + 
                                    float(temp_data.get("house_rent_allowance", 0) or 0) + 
                                    float(temp_data.get("overtime", 0) or 0) + 
                                    float(temp_data.get("arrear", 0) or 0) + 
                                    float(temp_data.get("incentive_pay", 0) or 0) + 
                                    float(temp_data.get("leave_encashment", 0) or 0) + 
                                    float(temp_data.get("travelling_allowance", 0) or 0) + 
                                    float(temp_data.get("abry_scheme", 0) or 0)    )
      temp_data['net_gross'] = ( float(temp_data.get("earned_gross", 0) or 0) - 
                                 float(temp_data.get("late_hours_deduction", 0) or 0)   )
      
      temp_data['total'] = ( float(temp_data.get("earned_gross", 0) or 0) -
                             float(temp_data.get("employee_contribution_pf", 0) or 0) - 
                             float(temp_data.get("employee_contribution_esi", 0) or 0) - 
                             float(temp_data.get("late_hours_deduction", 0) or 0) - 
                             float(temp_data.get("advance_deduction", 0) or 0) - 
                             float(temp_data.get("tax_deducted_at_source", 0) or 0) - 
                             float(temp_data.get("professional_tax", 0) or 0) - 
                             float(temp_data.get("income_tax", 0) or 0) - 
                             float(temp_data.get("gmi", 0) or 0)  )


def update_total_data(data, filters) :
    report_data = []
    temp_data_1 = {}
    temp_data_2 = {}

    hdfc_total = 0
    other_total = 0
    
    for row in data :
            if filters.get("based_on") == "Occupassion":
                if row.get("occupation", False) :
                    if row.get("occupation") == "Worker" :
                        update_data(row, 2, temp_data_2, "Worker")
                    elif row.get("occupation") == "Staff"  :
                        update_data(row, 1, temp_data_1, "Staff")
                    else :
                        update_data(row, 2, temp_data_2, "Worker")
                else :
                    update_data(row, 2, temp_data_2, "Worker")
            else :
                if row.get("occupation_accounts", False) :
                    if row.get("occupation_accounts") == "Direct" :
                        update_data(row, 1, temp_data_1, "Direct")
                    elif row.get("occupation_accounts") == "Indirect"  :
                        update_data(row, 2, temp_data_2, "Indirect")
                    else :
                        update_data(row, 2, temp_data_2, "Indirect")
                else :
                    update_data(row, 2, temp_data_2, "Indirect")


            if row.get("bank_name_new", False) :
                if row.get("bank_name_new") == "HDFC Bank" :
                    hdfc_total += float(row.get("net_pay", 0) or 0)
                else :
                    other_total += float(row.get("net_pay", 0) or 0)
            else :
                other_total += float(row.get("net_pay", 0) or 0)


    report_data.append(temp_data_1)
    report_data.append(temp_data_2)
    total_data = {}


    for key in temp_data_1:
        if key == "sr_no" :
             continue
        
        total_data[key] = temp_data_1.get(key, 0) + temp_data_2.get(key, 0)

        if key == "occupation" : 
            total_data[key] = "Total"
            # total_data[key] = "<b>Total</b>"
        


    report_data.append(total_data)

    grand_total = hdfc_total + other_total

    return {"report_data": report_data, "hdfc_total": hdfc_total, "other_total": other_total, "grand_total": grand_total}

     


def get_columns(filters):
     
	return [
        {"fieldname": "sr_no", "fieldtype": "Data", "label": "S.No", "width": 120},
        {"fieldname": "occupation", "fieldtype": "Data", "label": "Occupation", "width": 120} if filters.get("based_on") == "Occupassion" else {"fieldname": "occupation", "fieldtype": "Data", "label": "Account", "width": 120},
        {"fieldname": "gross", "fieldtype": "Currency", "label": "Gross Pay", "width": 140},
        {"fieldname": "total_employee", "fieldtype": "Data", "label": "Employee", "options": "Employee", "width": 120},
        {"fieldname": "basic", "fieldtype": "Currency", "label": "<span style='color: green'>Basic</span>", "width": 140,},
        {"fieldname": "city_compensatory_allowance", "fieldtype": "Currency", "label": "<span style='color: green'>City Compensatory Allowance</span>", "width": 210, "group": "Earnings"},
        {"fieldname": "washing_allowance", "fieldtype": "Currency", "label": "<span style='color: green'>Washing Allowance</span>", "width": 150, "group": "Earnings"},
        {"fieldname": "house_rent_allowance", "fieldtype": "Currency", "label": "<span style='color: green'>House Rent Allowance</span>", "width": 180, "group": "Earnings"},
        {"fieldname": "overtime", "fieldtype": "Currency", "label": "<span style='color: green'>Overtime</span>", "width": 120, "group": "Earnings"},
        {"fieldname": "arrear", "fieldtype": "Currency", "label": "<span style='color: green'>Arrear</span>", "width": 120, "group": "Earnings"},
        {"fieldname": "incentive_pay", "fieldtype": "Currency", "label": "<span style='color: green'>Incentive Pay</span>", "width": 130, "group": "Earnings"},
        {"fieldname": "leave_encashment", "fieldtype": "Currency", "label": "<span style='color: green'>Leave Encashment</span>", "width": 150, "group": "Earnings"},
        {"fieldname": "travelling_allowance", "fieldtype": "Currency", "label": "<span style='color: green'>Travelling Allowance</span>", "width": 150, "group": "Earnings"},
        {"fieldname": "abry_scheme", "fieldtype": "Currency", "label": "<span style='color: green'>ABRY Scheme</span>", "width": 120, "group": "Earnings"},
        {"fieldname": "earned_gross", "fieldtype": "Currency", "label": "Earned Gross", "width": 150, },
        {"fieldname": "net_gross", "fieldtype": "Currency", "label": "Net Gross", "width": 150, },
        {"fieldname": "employee_contribution_pf", "fieldtype": "Currency", "label": "<span style='color: red'>Employee Contribution PF</span>", "width": 200, "group": "Deductions"},
        {"fieldname": "employee_contribution_esi", "fieldtype": "Currency", "label": "<span style='color: red'>Employee Contribution ESI</span>", "width": 200, "group": "Deductions"},
        {"fieldname": "late_hours_deduction", "fieldtype": "Currency", "label": "<span style='color: red'>Late Hours Deduction</span>", "width": 160, "group": "Deductions"},
        {"fieldname": "advance_deduction", "fieldtype": "Currency", "label": "<span style='color: red'>Advance Deduction</span>", "width": 160, "group": "Deductions"},
        {"fieldname": "tax_deducted_at_source", "fieldtype": "Currency", "label": "<span style='color: red'>Tax Deducted At Source</span>", "width": 180, "group": "Deductions"},
        {"fieldname": "professional_tax", "fieldtype": "Currency", "label": "<span style='color: red'>Professional Tax</span>", "width": 140, "group": "Deductions"},
        {"fieldname": "income_tax", "fieldtype": "Currency", "label": "<span style='color: red'>Income Tax</span>", "width": 120, "group": "Deductions"},
        {"fieldname": "gmi", "fieldtype": "Currency", "label": "<span style='color: red'>GMI</span>", "width": 100, "group": "Deductions"},
        {"fieldname": "total", "fieldtype": "Currency", "label": "Total", "width": 150, }
    ]


def get_data(filters):
    conditions = get_conditions(filters)

    query = f"""
        SELECT 
            tss.occupation,
            tss.occupation_accounts,
            tss.gross_monthly_salary,
            tss.net_pay,
            tss.employee,
            tss.bank_name_new,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'B' ) basic,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'CCA' ) city_compensatory_allowance,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'WA' ) washing_allowance,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'HRA' ) house_rent_allowance,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'OT' ) overtime,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'A' ) arrear,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'IP' ) incentive_pay,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'LE' ) leave_encashment,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'TA' ) travelling_allowance,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'ABRY' ) abry_scheme,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'PF' ) employee_contribution_pf,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'ESI' ) employee_contribution_esi,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'LHD' ) late_hours_deduction,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'AD' ) advance_deduction,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'TDS' ) tax_deducted_at_source,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'PT' ) professional_tax,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'IT' ) income_tax,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'GMI' ) gmi
        FROM `tabSalary Slip` AS tss 
        WHERE tss.docstatus <= 1
              {conditions}
    	"""
    return frappe.db.sql(query, as_dict=1)
    

def get_conditions(filters):
    conditions = ""
    
    if filters.get("company"):
        conditions += " AND tss.company = '{}'".format(filters.get("company"))

    if filters.get("month"):
        month = [
            "January", 
            "February",
            "March",
            "April",
            "May", 
            "June", 
            "July", 
            "August", 
            "September", 
            "October", 
            "November", 
            "December"
        ].index(filters["month"]) + 1
        conditions += f" AND MONTH(tss.start_date) = {month}"
    
    if filters.get("year"):
        conditions += f" AND YEAR(tss.start_date) = {filters.get('year')}"
    
    return conditions


def get_chart_data(hdfc, other):
	# labels = ["HDFC Bank", "Other bank"]
	labels = ["Bank",]
	hdfc = [hdfc]
	other = [other]
	# overdue = []

	# for project in data:
	# 	labels.append(project.project_name)
	# 	total.append(project.total_tasks)
	# 	completed.append(project.completed_tasks)
	# 	overdue.append(project.overdue_tasks)

	return {
		"data": {
			"labels": labels[:2],
			"datasets": [
				{"name": ("HDFC Bank"), "values": hdfc[:2]},
				{"name": ("Other bank"), "values": other[:2]},
			],
		},
		"type": "bar",
		"colors": ["#F1A069", "#89d3f5",],
		"barOptions": {"stacked": True},
	}



# def get_chart_data(data):
# 	labels = []
# 	total = []
# 	completed = []
# 	overdue = []

# 	for project in data:
# 		labels.append(project.project_name)
# 		total.append(project.total_tasks)
# 		completed.append(project.completed_tasks)
# 		overdue.append(project.overdue_tasks)

# 	return {
# 		"data": {
# 			"labels": labels[:30],
# 			"datasets": [
# 				{"name": _("Overdue"), "values": overdue[:30]},
# 				{"name": _("Completed"), "values": completed[:30]},
# 				{"name": _("Total Tasks"), "values": total[:30]},
# 			],
# 		},
# 		"type": "bar",
# 		"colors": ["#fc4f51", "#78d6ff", "#7575ff"],
# 		"barOptions": {"stacked": True},
# 	}



def get_report_summary(hdfc_total, other_total, grand_total):

	return [
		{
			"value": hdfc_total,
			"indicator": "Blue" ,
			"label": "HDFC Bank",
			"datatype": "Currency",
		},
		{
			"value": other_total,
			"indicator": "Blue",
			"label": "Other bank",
			"datatype": "Currency",
		},
		{
			"value": grand_total,
			"indicator": "Green",
			"label": "Total",
			"datatype": "Currency",
		},
	]



@frappe.whitelist()
def get_excel_data(filters):
	filters = json.loads(filters)
	# report_data = json.loads(data)
	file_name = "Salary Sheet"
	if filters.get("company"):
		file_name  += " - "+ filters.get("company") 
	if filters.get("month"):
		file_name  += " - "+ filters.get("month") 
	if filters.get("year"):
		file_name  += " - "+ filters.get("year") 
	data = ""

	return {
        "content": data,
        "filename": file_name,  
        "extension": 'xlsx',
        "filters":filters
    }


@frappe.whitelist()
def download_file():    

    data = frappe._dict(frappe.local.form_dict)
    filters =  json.loads(data["filters"])

    company = ""
    if filters.get("company"):
        company = filters.get("company") 

	# workbook = openpyxl.Workbook(write_only=True)
    workbook = openpyxl.Workbook()
    current_sheet = workbook.active
    border_style_thin = Side(border_style='thin')
    border_style_thick = Side(border_style='thick')
    bold_font = Font(bold=True)

	# current_sheet.title = "Salary Sheet"

	# first row
     
    current_sheet.row_dimensions[1].height = 30
    current_sheet.merge_cells("A1:Y1")
    update_border_font_align(current_sheet, 'A1', company, border_style_thick)
    current_sheet['A1'].font = Font(size=16, bold=True)


    current_sheet.row_dimensions[2].height = 25
    current_sheet.merge_cells("A2:D2")
    update_border_font_align(current_sheet, 'A2', "", border_style_thin)
    # current_sheet['A2'].border = set_border(border_style_thick)
    # current_sheet['D2'].border = set_border(border_style_thick)
    current_sheet.merge_cells("E2:N2")
    current_sheet["E2"].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
    update_border_font_align(current_sheet, 'E2', "Earning", border_style_thin); 
    current_sheet['E2'].font = Font(size=14, bold=True)
	# current_sheet['E2'] = "Earning"
	# current_sheet['E2'].font = bold_font
	# current_sheet['E2'].alignment = Alignment(horizontal='center', vertical='center')
	# current_sheet['E2'].border = set_border(border_style_thick)
    # current_sheet['N2'].border = set_border(border_style_thick)
    current_sheet.merge_cells("O2:P2")
    update_border_font_align(current_sheet, 'O2', "", border_style_thin)
    # current_sheet['O2'].border = set_border(border_style_thick)
    # current_sheet['P2'].border = set_border(border_style_thick)

    current_sheet.merge_cells("Q2:W2")
    current_sheet["Q2"].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
    update_border_font_align(current_sheet, 'Q2', "Deduction", border_style_thin) ; 
    current_sheet['Q2'].font = Font(size=14, bold=True)
	# current_sheet['Q2'] = "Deduction"
	
	# current_sheet['Q2'].alignment = Alignment(horizontal='center', vertical='center')
	# current_sheet['Q2'].border = set_border(border_style_thick)
    # current_sheet['W2'].border = set_border(border_style_thick)
    
    current_sheet.merge_cells("X2:Y2")
    update_border_font_align(current_sheet, 'X2', "", border_style_thin)
	# current_sheet['Y2'] = "Total"
	# current_sheet['Y2'].font = bold_font
	# current_sheet['Y2'].border = set_border(border_style_thick)

	# second row
	# 
    current_sheet.row_dimensions[3].height = 50
    current_sheet.column_dimensions['B'].width = 18
    current_sheet.column_dimensions['C'].width = 15
    current_sheet.column_dimensions['D'].width = 15
    current_sheet.column_dimensions['E'].width = 15
    current_sheet.column_dimensions['F'].width = 15
    current_sheet.column_dimensions['G'].width = 15
    current_sheet.column_dimensions['H'].width = 15
    current_sheet.column_dimensions['I'].width = 15
    current_sheet.column_dimensions['K'].width = 15
    current_sheet.column_dimensions['L'].width = 15
    current_sheet.column_dimensions['M'].width = 15
    current_sheet.column_dimensions['N'].width = 15
    current_sheet.column_dimensions['O'].width = 15
    current_sheet.column_dimensions['P'].width = 15
    current_sheet.column_dimensions['Q'].width = 15
    current_sheet.column_dimensions['R'].width = 15
    current_sheet.column_dimensions['S'].width = 15
    current_sheet.column_dimensions['T'].width = 15
    current_sheet.column_dimensions['U'].width = 15
    current_sheet.column_dimensions['V'].width = 15
    current_sheet.column_dimensions['W'].width = 15
    current_sheet.column_dimensions['X'].width = 15
    current_sheet.column_dimensions['Y'].width = 15

    update_border_font_align(current_sheet, 'A3', "S.No.", border_style_thin)
    if filters.get("based_on") == "Occupassion" :
        update_border_font_align(current_sheet, 'B3', "Occupation", border_style_thin)
    else :
        update_border_font_align(current_sheet, 'B3', "Account", border_style_thin)
        
    update_border_font_align(current_sheet, 'C3', "Gross", border_style_thin)
    update_border_font_align(current_sheet, 'D3', "Employee", border_style_thin)
    update_border_font_align(current_sheet, 'E3', "Basic", border_style_thin)
    update_border_font_align(current_sheet, 'F3', "City Compensatory Allowance", border_style_thin)
    update_border_font_align(current_sheet, 'G3', "Washing Allowance", border_style_thin)
    update_border_font_align(current_sheet, 'H3', "House Rent Allowance	", border_style_thin)
    update_border_font_align(current_sheet, 'I3', "Overtime", border_style_thin)
    update_border_font_align(current_sheet, 'J3', "Arrear", border_style_thin)
    update_border_font_align(current_sheet, 'K3', "Incentive Pay", border_style_thin)
    update_border_font_align(current_sheet, 'L3', "Leave Encashment", border_style_thin)
    update_border_font_align(current_sheet, 'M3', "Travelling Allowance", border_style_thin)
    update_border_font_align(current_sheet, 'N3', "ABRY Scheme", border_style_thin)
    update_border_font_align(current_sheet, 'O3', "Earned Gross", border_style_thin)
    update_border_font_align(current_sheet, 'P3', "Net Gross", border_style_thin)
    update_border_font_align(current_sheet, 'Q3', "Employee Contribution PF", border_style_thin)
    update_border_font_align(current_sheet, 'R3', "Employee Contribution ESI", border_style_thin)
    update_border_font_align(current_sheet, 'S3', "Late Hours Deduction", border_style_thin)
    update_border_font_align(current_sheet, 'T3', "Advance Deduction", border_style_thin)
    update_border_font_align(current_sheet, 'U3', "Tax Deducted at Source", border_style_thin)
    update_border_font_align(current_sheet, 'V3', "Professional Tax", border_style_thin)
    update_border_font_align(current_sheet, 'W3', "Income Tax", border_style_thin)
    update_border_font_align(current_sheet, 'X3', "GMI", border_style_thin)
    update_border_font_align(current_sheet, 'Y3', "Total", border_style_thin)
	# current_sheet['A3'] = "S.No."
	# current_sheet['A3'].font = bold_font
	# current_sheet['A3'].border = set_border(border_style_thin)


    query_data = get_data(filters)
    # frappe.response["filters"] = data["filters"]
    # frappe.response["filters1"] = json.loads(data["filters"])

    temp_data = update_total_data(query_data, filters)
    report_data = temp_data['report_data']
    hdfc_total = temp_data['hdfc_total']
    other_total = temp_data['other_total']
    grand_total = temp_data['grand_total']

    row_no = 4
    for row in report_data :
        update_border_font_align(current_sheet, f"A{row_no}", row.get("sr_no", ""), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"B{row_no}", row.get("occupation", ""), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"C{row_no}", format_number(row.get("gross", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"D{row_no}", row.get("total_employee", ""), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"E{row_no}", format_number(row.get("basic", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"F{row_no}", format_number(row.get("city_compensatory_allowance", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"G{row_no}", format_number(row.get("washing_allowance", "")), border_style_thin , 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"H{row_no}", format_number(row.get("house_rent_allowance", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"I{row_no}", format_number(row.get("overtime", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"J{row_no}", format_number(row.get("arrear", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"K{row_no}", format_number(row.get("incentive_pay", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"L{row_no}", format_number(row.get("leave_encashment", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"M{row_no}", format_number(row.get("travelling_allowance", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"N{row_no}", format_number(row.get("abry_scheme", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"O{row_no}", format_number(row.get("earned_gross", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"P{row_no}", format_number(row.get("net_gross", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"Q{row_no}", format_number(row.get("employee_contribution_pf", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"R{row_no}", format_number(row.get("employee_contribution_esi", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"S{row_no}", format_number(row.get("late_hours_deduction", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"T{row_no}", format_number(row.get("advance_deduction", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"U{row_no}", format_number(row.get("tax_deducted_at_source", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"V{row_no}", format_number(row.get("professional_tax", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"W{row_no}", format_number(row.get("income_tax", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"X{row_no}", format_number(row.get("gmi", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        update_border_font_align(current_sheet, f"Y{row_no}", format_number(row.get("total", "")), border_style_thin, 'yes' if row_no ==6 else 'no' )
        row_no += 1
    
    update_border_font_align(current_sheet, 'B9', "HDFC Bank", border_style_thin)
    update_border_font_align(current_sheet, 'C9', format_number(hdfc_total), border_style_thin)
    update_border_font_align(current_sheet, 'B10', "Other Bank", border_style_thin)
    update_border_font_align(current_sheet, 'C10', format_number(other_total), border_style_thin)
    update_border_font_align(current_sheet, 'B11', "Total", border_style_thin)
    update_border_font_align(current_sheet, 'C11', format_number(grand_total), border_style_thin)

	


    byte_file = io.BytesIO()
    workbook.save(byte_file)

    frappe.response["type"] = "binary"
    frappe.response["filecontent"] = byte_file.getvalue()
    frappe.response["filename"] = f"{data['filename']}.{data['extension']}"


def set_border(border_type):
	return Border(top=border_type, left=border_type, right=border_type, bottom=border_type)

def update_border_font_align(sheet, index, name, border_type, bold='yes') :
    sheet[index] = name
    if bold == 'yes':
        sheet[index].font = Font(bold=True)
    sheet[index].border = set_border(border_type)
    # sheet[index].alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
    sheet[index].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # sheet[index].width = 200
    return sheet



	# data = frappe._dict(frappe.local.form_dict)
	# frappe.response["filename"] = (
    #     frappe.scrub(f"{data['report_name']} {data['company']}") + ".txt"
    # )
	# frappe.response["filecontent"] = data["data"]
	# frappe.response["content_type"] = "application/txt"
	# frappe.response["type"] = "download"


def format_number(value) :
     if float(value or 0):
        #   return "{:,.0f}0".format(value)
        #   return frappe.utils.fmt_money(value, precision=0, currency='INR')
          return frappe.utils.fmt_money(value, precision=0)
     else :
          return 0
      