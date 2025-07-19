# Copyright (c) 2025, pinkcity and contributors
# For license information, please see license.txt

import frappe, openpyxl, json, io
import openpyxl.styles
import openpyxl.workbook
from openpyxl.styles import Border, Side, Alignment, Font
from frappe.utils import flt


def execute(filters):
    # headings = get_heading()
    columns = get_columns()
    data = get_data(filters)
    report_data = []
    temp_data_1 = {}
    temp_data_2 = {}

    hdfc_total = 0
    other_total = 0
    
    # report_data[0] = []
    # report_data[1] = []
    for row in data :
            if row.get("occupation", False) :
                if row.get("occupation") == "Worker" :
                    update_data(row, 2, temp_data_2, "Worker")
                elif row.get("occupation") == "Staff"  :
                    update_data(row, 1, temp_data_1, "Staff")
                else :
                    update_data(row, 2, temp_data_2, "Worker")
            else :
                update_data(row, 2, temp_data_2, "Worker")


            if row.get("bank_name_new", False) :
                if row.get("bank_name_new") == "HDFC Bank" :
                    hdfc_total += float(row.get("net_pay", 0) or 0)
                else :
                    other_total += float(row.get("net_pay", 0) or 0)
            else :
                other_total += float(row.get("net_pay", 0) or 0)
    report_data.append(temp_data_1)
    report_data.append(temp_data_2)
    total_data = {}
    for key in temp_data_1:
        total_data[key] = temp_data_1.get(key, 0) + temp_data_2.get(key, 0)
    report_data.append(total_data)
    # if("bank_name_new" == "HDFC Bank"):
    #     hdfc_total = sum(bank_name_new)
    # else :
    #     other_total = sum(bank_name_new)
    # hdfc_total = sum("bank_name_new" == "HDFC Bank")
    # other_total = sum("bank_name_new" != "HDFC Bank")
    grand_total = hdfc_total + other_total
    report_data.append({
		"total_employee": "<p style='font-weight: bold;'>HDFC BANK</p>",
		"washing_allowance": round(hdfc_total, 2)
	})
    report_data.append({
		"total_employee": "<p style='font-weight: bold;'>OTHER BANK</p>",
		"washing_allowance": round(other_total, 2)
	})
    report_data.append({
		"total_employee": "<p style='font-weight: bold;'>GRAND TOTAL</p>",
		"washing_allowance": round(grand_total, 2)
	})
    return columns, report_data

def update_data(data_value, sr_no, temp_data, occupation) :
      temp_data['sr_no'] = sr_no
      temp_data['occupation'] = occupation
      temp_data['gross'] = float(temp_data.get("gross", 0) or 0) + float(data_value.get("gross_monthly_salary", 0) or 0)
      temp_data['total_employee'] = float(temp_data.get("total_employee", 0) or 0) + 1
      temp_data['basic'] = float(temp_data.get("basic", 0) or 0) + float(data_value.get("basic", 0) or 0)
      temp_data['city_compensatory_allowance'] = float(temp_data.get("city_compensatory_allowance", 0) or 0) + float(data_value.get("city_compensatory_allowance", 0) or 0)
      temp_data['washing_allowance'] = float(temp_data.get("washing_allowance", 0) or 0) + float(data_value.get("washing_allowance", 0) or 0)
      temp_data['house_rent_allowance'] = float(temp_data.get("house_rent_allowance", 0) or 0) + float(data_value.get("house_rent_allowance", 0) or 0)
      temp_data['overtime'] = float(temp_data.get("overtime", 0) or 0) + float(data_value.get("overtime", 0) or 0)
      temp_data['arrear'] = float(temp_data.get("arrear", 0) or 0) + float(data_value.get("arrear", 0) or 0)
      temp_data['incentive_pay'] = float(temp_data.get("incentive_pay", 0) or 0) + float(data_value.get("incentive_pay", 0) or 0)
      temp_data['leave_encashment'] = float(temp_data.get("leave_encashment", 0) or 0) + float(data_value.get("leave_encashment", 0) or 0)
      temp_data['travelling_allowance'] = float(temp_data.get("travelling_allowance", 0) or 0) + float(data_value.get("travelling_allowance", 0) or 0)
      temp_data['abry_scheme'] = float(temp_data.get("abry_scheme", 0) or 0) + float(data_value.get("abry_scheme", 0) or 0)
      temp_data['earned_gross'] = float(temp_data.get("earned_gross", 0) or 0) + float(data_value.get("earned_gross", 0) or 0)
      temp_data['net_gross'] = float(temp_data.get("net_gross", 0) or 0) + float(data_value.get("net_gross", 0) or 0)
      temp_data['employee_contribution_pf'] = float(temp_data.get("employee_contribution_pf", 0) or 0) + float(data_value.get("employee_contribution_pf", 0) or 0)
      temp_data['employee_contribution_esi'] = float(temp_data.get("employee_contribution_esi", 0) or 0) + float(data_value.get("employee_contribution_esi", 0) or 0)
      temp_data['late_hours_deduction'] = float(temp_data.get("late_hours_deduction", 0) or 0) + float(data_value.get("late_hours_deduction", 0) or 0)
      temp_data['advance_deduction'] = float(temp_data.get("advance_deduction", 0) or 0) + float(data_value.get("advance_deduction", 0) or 0)
      temp_data['tax_deducted_at_source'] = float(temp_data.get("tax_deducted_at_source", 0) or 0) + float(data_value.get("tax_deducted_at_source", 0) or 0)
      temp_data['income_tax'] = float(temp_data.get("income_tax", 0) or 0) + float(data_value.get("income_tax", 0) or 0)
      temp_data['gmi'] = float(temp_data.get("gmi", 0) or 0) + float(data_value.get("gmi", 0) or 0)
      temp_data['hdfc_total'] = float(temp_data.get("hdfc_total", 0) or 0) + float(data_value.get("hdfc_total", 0) or 0)
      temp_data['other_total'] = float(temp_data.get("other_total", 0) or 0) + float(data_value.get("other_total", 0) or 0)
      temp_data['grand_total'] = float(temp_data.get("grand_total", 0) or 0) + float(data_value.get("grand_total", 0) or 0)



# def get_data(filters):
# 	conditions = get_conditions(filters)

# 	query = f"""
# 				SELECT 
# 					tss.occupation,
# 					tss.company,
# 					SUM(tss.gross_pay) AS gross,
# 					COUNT(tss.employee_name) AS employee,
# 					SUM(tsd.amount) AS basic,
# 					SUM(CASE WHEN tsd.salary_component = 'City Compensatory Allowance' THEN tsd.amount ELSE 0 END) AS city_compensatory_allowance,
# 					SUM(CASE WHEN tsd.salary_component = 'Washing Allowance' THEN tsd.amount ELSE 0 END) AS washing_allowance,
# 					SUM(CASE WHEN tsd.salary_component = 'House Rent Allowance' THEN tsd.amount ELSE 0 END) AS house_rent_allowance,
# 					SUM(CASE WHEN tsd.salary_component = 'OT' THEN tsd.amount ELSE 0 END) AS overtime,
# 					SUM(CASE WHEN tsd.salary_component = 'Arrear' THEN tsd.amount ELSE 0 END) AS arrear,
# 					SUM(CASE WHEN tsd.salary_component = 'Incentive Pay' THEN tsd.amount ELSE 0 END) AS incentive_pay,
# 					SUM(CASE WHEN tsd.salary_component = 'Leave Encashment' THEN tsd.amount ELSE 0 END) AS leave_encashment,
# 					SUM(CASE WHEN tsd.salary_component = 'Travelling Allowance' THEN tsd.amount ELSE 0 END) AS travelling_allowance,
# 					SUM(CASE WHEN tsd.salary_component = 'ABRY Scheme' THEN tsd.amount ELSE 0 END) AS abry_scheme,
# 					SUM(CASE WHEN tsd.salary_component = 'Earned Gross' THEN tsd.amount ELSE 0 END) AS earned_gross,
# 					SUM(CASE WHEN tsd.salary_component = 'Net Gross' THEN tsd.amount ELSE 0 END) AS net_gross,
# 					SUM(CASE WHEN tsd.salary_component = 'Employee Contribution PF' THEN tsd.amount ELSE 0 END) AS employee_contribution_pf,
# 					SUM(CASE WHEN tsd.salary_component = 'Employee Contribution ESI' THEN tsd.amount ELSE 0 END) AS employee_contribution_esi,
# 					SUM(CASE WHEN tsd.salary_component = 'Late Hour Deduction' THEN tsd.amount ELSE 0 END) AS late_hours_deduction,
# 					SUM(CASE WHEN tsd.salary_component = 'Advance Deduction' THEN tsd.amount ELSE 0 END) AS advance_deduction,
# 					SUM(CASE WHEN tsd.salary_component = 'Tax Deducted at Source' THEN tsd.amount ELSE 0 END) AS tax_deducted_at_source,
# 					SUM(CASE WHEN tsd.salary_component = 'Professional Tax' THEN tsd.amount ELSE 0 END) AS professional_tax,
# 					SUM(CASE WHEN tsd.salary_component = 'Income Tax' THEN tsd.amount ELSE 0 END) AS income_tax,
# 					SUM(CASE WHEN tsd.salary_component = 'GMI' THEN tsd.amount ELSE 0 END) AS gmi,
# 					SUM(CASE WHEN tsd.salary_component = 'Total' THEN tsd.amount ELSE 0 END) AS total
					
# 				FROM `tabSalary Component` AS tsc
# 				LEFT JOIN `tabSalary Detail` AS tsd 
# 					ON tsd.abbr = tsc.salary_component_abbr
# 				LEFT JOIN `tabSalary Slip` AS tss 
# 					ON tss.name = tsd.parent
# 				WHERE tss.occupation IN ('Worker', 'Staff')  
# 				AND tss.docstatus <= 1                     
# 				GROUP BY 
# 					tss.occupation
# 				ORDER BY 
# 					tss.occupation;


# 				{conditions}
# 				"""


# 	return frappe.db.sql(query, as_dict=1,)



def get_columns():
	return [
        {"fieldname": "sr_no", "fieldtype": "Data", "label": "S.No", "width": 120},
        {"fieldname": "occupation", "fieldtype": "Data", "label": "Occupation", "width": 120},
        {"fieldname": "gross", "fieldtype": "Data", "label": "Gross Pay", "width": 120},
        {"fieldname": "total_employee", "fieldtype": "Data", "label": "Employee", "options": "Employee", "width": 120},
        {"fieldname": "basic", "fieldtype": "Data", "label": "Basic", "width": 120, "group": "Earnings"},
        {"fieldname": "city_compensatory_allowance", "fieldtype": "Data", "label": "City Compensatory Allowance", "width": 150, "group": "Earnings"},
        {"fieldname": "washing_allowance", "fieldtype": "Data", "label": "Washing Allowance", "width": 120, "group": "Earnings"},
        {"fieldname": "house_rent_allowance", "fieldtype": "Data", "label": "House Rent Allowance", "width": 150, "group": "Earnings"},
        {"fieldname": "overtime", "fieldtype": "Data", "label": "Overtime", "width": 120, "group": "Earnings"},
        {"fieldname": "arrear", "fieldtype": "Data", "label": "Arrear", "width": 120, "group": "Earnings"},
        {"fieldname": "incentive_pay", "fieldtype": "Data", "label": "Incentive Pay", "width": 120, "group": "Earnings"},
        {"fieldname": "leave_encashment", "fieldtype": "Data", "label": "Leave Encashment", "width": 140, "group": "Earnings"},
        {"fieldname": "travelling_allowance", "fieldtype": "Data", "label": "Travelling Allowance", "width": 140, "group": "Earnings"},
        {"fieldname": "abry_scheme", "fieldtype": "Data", "label": "Abry Scheme", "width": 120, "group": "Earnings"},
        {"fieldname": "earned_gross", "fieldtype": "Data", "label": "Earned Gross", "width": 120, "group": "Earnings"},
        {"fieldname": "net_gross", "fieldtype": "Data", "label": "Net Gross", "width": 120, "group": "Earnings"},
        {"fieldname": "employee_contribution_pf", "fieldtype": "Data", "label": "Employee Contribution Pf", "width": 160, "group": "Deductions"},
        {"fieldname": "employee_contribution_esi", "fieldtype": "Data", "label": "Employee Contribution Esi", "width": 160, "group": "Deductions"},
        {"fieldname": "late_hours_deduction", "fieldtype": "Data", "label": "Late Hours Deduction", "width": 140, "group": "Deductions"},
        {"fieldname": "advance_deduction", "fieldtype": "Data", "label": "Advance Deduction", "width": 140, "group": "Deductions"},
        {"fieldname": "tax_deducted_at_source", "fieldtype": "Data", "label": "Tax Deducted At Source", "width": 160, "group": "Deductions"},
        {"fieldname": "professional_tax", "fieldtype": "Data", "label": "Professional Tax", "width": 140, "group": "Deductions"},
        {"fieldname": "income_tax", "fieldtype": "Data", "label": "Income Tax", "width": 120, "group": "Deductions"},
        {"fieldname": "gmi", "fieldtype": "Data", "label": "Gmi", "width": 100, "group": "Deductions"},
        {"fieldname": "total", "fieldtype": "Data", "label": "Total", "width": 120, "group": "Deductions"}
    ]

# def get_conditions(filters):
# 	conditions = ""
	
# 	if filters.get("company"):
# 		conditions += " AND tss.company = '" + filters.get("company") +"' " 

# 	if filters.get("month"):
# 		month = [
# 			"January",
# 			"Febuary",
# 			"March",
# 			"April",
# 			"May",
# 			"June",
# 			"July",
# 			"August",
# 			"September",
# 			"October",
# 			"November",
# 			"December",
# 		].index(filters["month"]) + 1
# 		conditions += f" AND MONTH(tss.posting_date) = {month} "
	
# 	if filters.get("year"):
# 		conditions += f" AND YEAR(tss.posting_date) =  " + filters.get("year") +" " 
	
# 	if filters.get("employee"):
# 		conditions += f" AND tss.employee =  '" + filters.get("employee") +"' " 
	

# 	return conditions


def get_data(filters):
    conditions = get_conditions(filters)

    query = f"""
        SELECT 
            tss.occupation,
            tss.gross_monthly_salary,
            tss.net_pay,
            tss.employee,
            tss.bank_name_new,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'B' ) basic,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'CCA' ) city_compensatory_allowance,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'WA' ) washing_allowance,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'HRA' ) house_rent_allowance,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'OT' ) overtime,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'A' ) arrear,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'IP' ) incentive_pay,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'LE' ) leave_encashment,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'TA' ) travelling_allowance,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'ABRY' ) abry_scheme,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'PF' ) employee_contribution_pf,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'ESI' ) employee_contribution_esi,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'LHD' ) late_hours_deduction,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'AD' ) advance_deduction,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'TDS' ) tax_deducted_at_source,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'PT' ) professional_tax,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'IT' ) income_tax,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'GMI' ) gmi
        FROM `tabSalary Slip` AS tss 
        WHERE tss.docstatus <= 1
        {conditions};
    	"""
    return frappe.db.sql(query, as_dict=1)
    # Create the query with conditions directly embedded
    # query = f"""
    #     SELECT 
    #         tss.occupation,
    #         tss.gross_monthly_salary,
    #         tss.employee,
    #         ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'B' ) basic
    #     FROM `tabSalary Slip` AS tss 
    #     WHERE tss.docstatus <= 1
    #     {conditions};
    # 	"""

        # ORDER BY 
        #     tss.occupation
    
	# ,
    #         SUM(tsd.amount) AS basic,
    #         SUM(CASE WHEN tsd.salary_component = 'City Compensatory Allowance' THEN tsd.amount ELSE 0 END) AS city_compensatory_allowance,
    #         SUM(CASE WHEN tsd.salary_component = 'Washing Allowance' THEN tsd.amount ELSE 0 END) AS washing_allowance,
    #         SUM(CASE WHEN tsd.salary_component = 'House Rent Allowance' THEN tsd.amount ELSE 0 END) AS house_rent_allowance,
    #         SUM(CASE WHEN tsd.salary_component = 'OT' THEN tsd.amount ELSE 0 END) AS overtime,
    #         SUM(CASE WHEN tsd.salary_component = 'Arrear' THEN tsd.amount ELSE 0 END) AS arrear,
    #         SUM(CASE WHEN tsd.salary_component = 'Incentive Pay' THEN tsd.amount ELSE 0 END) AS incentive_pay,
    #         SUM(CASE WHEN tsd.salary_component = 'Leave Encashment' THEN tsd.amount ELSE 0 END) AS leave_encashment,
    #         SUM(CASE WHEN tsd.salary_component = 'Travelling Allowance' THEN tsd.amount ELSE 0 END) AS travelling_allowance,
    #         SUM(CASE WHEN tsd.salary_component = 'ABRY Scheme' THEN tsd.amount ELSE 0 END) AS abry_scheme,
    #         SUM(CASE WHEN tsd.salary_component = 'Earned Gross' THEN tsd.amount ELSE 0 END) AS earned_gross,
    #         SUM(CASE WHEN tsd.salary_component = 'Net Gross' THEN tsd.amount ELSE 0 END) AS net_gross,
    #         SUM(CASE WHEN tsd.salary_component = 'Employee Contribution PF' THEN tsd.amount ELSE 0 END) AS employee_contribution_pf,
    #         SUM(CASE WHEN tsd.salary_component = 'Employee Contribution ESI' THEN tsd.amount ELSE 0 END) AS employee_contribution_esi,
    #         SUM(CASE WHEN tsd.salary_component = 'Late Hour Deduction' THEN tsd.amount ELSE 0 END) AS late_hours_deduction,
    #         SUM(CASE WHEN tsd.salary_component = 'Advance Deduction' THEN tsd.amount ELSE 0 END) AS advance_deduction,
    #         SUM(CASE WHEN tsd.salary_component = 'Tax Deducted at Source' THEN tsd.amount ELSE 0 END) AS tax_deducted_at_source,
    #         SUM(CASE WHEN tsd.salary_component = 'Professional Tax' THEN tsd.amount ELSE 0 END) AS professional_tax,
    #         SUM(CASE WHEN tsd.salary_component = 'Income Tax' THEN tsd.amount ELSE 0 END) AS income_tax,
    #         SUM(CASE WHEN tsd.salary_component = 'GMI' THEN tsd.amount ELSE 0 END) AS gmi,
    #         SUM(CASE WHEN tsd.salary_component = 'Total' THEN tsd.amount ELSE 0 END) AS total

	# FROM `tabSalary Detail` AS tsd 
    #     LEFT JOIN `tabSalary Slip` AS tss ON tss.name = tsd.parent
    
    # Execute the query directly without using params
    

def get_conditions(filters):
    conditions = ""
    
    if filters.get("company"):
        conditions += " AND tss.company = '{}'".format(filters.get("company"))

    if filters.get("month"):
        month = [
            "January", 
            "February",
            "March",
            "April",
            "May", 
            "June", 
            "July", 
            "August", 
            "September", 
            "October", 
            "November", 
            "December"
        ].index(filters["month"]) + 1
        conditions += f" AND MONTH(tss.start_date) = {month}"
    
    if filters.get("year"):
        conditions += f" AND YEAR(tss.start_date) = {filters.get('year')}"
    
    
    return conditions


@frappe.whitelist()
def get_excel_data(filters):
	filters = json.loads(filters)
	# report_data = json.loads(data)
	file_name = "Salary Sheet"
	if filters.get("company"):
		file_name  += " - "+ filters.get("company") 
	data = ""

	return {
        "content": data,
        "filename": file_name,  
        "extension": 'xlsx'
    }


@frappe.whitelist()
def download_file():
	# workbook = openpyxl.Workbook(write_only=True)
	workbook = openpyxl.Workbook()
	current_sheet = workbook.active
	border_style_thin = Side(border_style='thin')
	border_style_thick = Side(border_style='thick')
	bold_font = Font(bold=True)

	# current_sheet.title = "Salary Sheet"

	# first row
	current_sheet.merge_cells("A1:D1")
	current_sheet['A1'].border = set_border(border_style_thick)
	current_sheet['D1'].border = set_border(border_style_thick)
	current_sheet.merge_cells("E1:N1")
	current_sheet['E1'] = "Earning"
	current_sheet['E1'].font = bold_font
	current_sheet['E1'].alignment = Alignment(horizontal='center', vertical='center')
	current_sheet['E1'].border = set_border(border_style_thick)
	current_sheet['N1'].border = set_border(border_style_thick)
	current_sheet.merge_cells("O1:P1")
	current_sheet['O1'].border = set_border(border_style_thick)
	current_sheet['P1'].border = set_border(border_style_thick)
	current_sheet.merge_cells("Q1:X1")
	current_sheet['Q1'] = "Deduction"
	current_sheet['Q1'].font = bold_font
	current_sheet['Q1'].alignment = Alignment(horizontal='center', vertical='center')
	current_sheet['Q1'].border = set_border(border_style_thick)
	current_sheet['X1'].border = set_border(border_style_thick)
	current_sheet['Y1'] = "Total"
	current_sheet['Y1'].font = bold_font
	current_sheet['Y1'].border = set_border(border_style_thick)

	# second row
	current_sheet['A2'] = "S.No."
	current_sheet['A2'].font = bold_font
	current_sheet['A2'].border = set_border(border_style_thin)
	


	byte_file = io.BytesIO()
	workbook.save(byte_file)

	data = frappe._dict(frappe.local.form_dict)
	frappe.response["type"] = "binary"
	frappe.response["filecontent"] = byte_file.getvalue()
	frappe.response["filename"] = f"{data['filename']}.{data['extension']}"


def set_border(border_type):
	return Border(top=border_type, left=border_type, right=border_type, bottom=border_type)

	# data = frappe._dict(frappe.local.form_dict)
	# frappe.response["filename"] = (
    #     frappe.scrub(f"{data['report_name']} {data['company']}") + ".txt"
    # )
	# frappe.response["filecontent"] = data["data"]
	# frappe.response["content_type"] = "application/txt"
	# frappe.response["type"] = "download"


      